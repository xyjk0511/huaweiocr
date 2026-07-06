import importlib
import inspect
import io
import json
import math
import os
import queue
import sys
import tempfile
import time
import threading
import types
import unittest
import subprocess
from unittest import mock

import numpy as np


def tearDownModule():
    for name in (
        "app_paths",
        "huaweiocr.io.paths_runtime",
        "barcode",
        "huaweiocr.barcode.generic",
        "crop",
        "cv2",
        "gui_app",
        "gui_app_en",
        "gui_pipeline",
        "numpy",
        "ocr",
        "huaweiocr.detect.ocr_engine",
        "paddle",
        "paddleocr",
        "pyzbar",
        "pyzbar.pyzbar",
        "scan2",
        "win_subprocess",
        "huaweiocr.io.win_subprocess",
    ):
        sys.modules.pop(name, None)


def _install_crop_import_fakes():
    cv2 = types.ModuleType("cv2")
    cv2.IMREAD_COLOR = 1
    cv2.IMWRITE_JPEG_QUALITY = 1
    cv2.imdecode = lambda *args, **kwargs: None
    cv2.imencode = lambda *args, **kwargs: (False, b"")
    cv2.dnn = types.SimpleNamespace(NMSBoxes=lambda *args, **kwargs: [])
    sys.modules["cv2"] = cv2

    numpy = types.ModuleType("numpy")
    numpy.uint8 = object()
    numpy.ceil = math.ceil
    numpy.tan = math.tan
    numpy.deg2rad = math.radians
    numpy.fromfile = lambda *args, **kwargs: b""
    sys.modules["numpy"] = numpy

    inference_sdk = types.ModuleType("inference_sdk")

    class DummyInferenceHTTPClient:
        def __init__(self, *args, **kwargs):
            pass

        def infer(self, *args, **kwargs):
            return {"predictions": []}

    inference_sdk.InferenceHTTPClient = DummyInferenceHTTPClient
    sys.modules["inference_sdk"] = inference_sdk


def _install_scan2_import_fakes():
    cv2 = types.ModuleType("cv2")
    cv2.IMREAD_COLOR = 1
    cv2.THRESH_BINARY = 0
    cv2.THRESH_OTSU = 0
    cv2.COLOR_BGR2GRAY = 1
    cv2.COLOR_GRAY2BGR = 2
    cv2.BORDER_CONSTANT = 0
    cv2.ROTATE_90_COUNTERCLOCKWISE = 0
    cv2.INTER_CUBIC = 0
    cv2.imread = lambda *args, **kwargs: None
    sys.modules["cv2"] = cv2

    numpy = types.ModuleType("numpy")
    sys.modules["numpy"] = numpy

    ocr = types.ModuleType("ocr")
    ocr.init_ocr = lambda: object()
    ocr.ocr_one_image = lambda *args, **kwargs: ([], "")
    sys.modules["ocr"] = ocr

    barcode = types.ModuleType("barcode")
    barcode.decode_small_patch = lambda *args, **kwargs: {"results": []}
    sys.modules["barcode"] = barcode

    app_paths = types.ModuleType("app_paths")
    user_data_dir = tempfile.mkdtemp(prefix="huaweiocr_test_data_")
    app_paths.ensure_models_installed = lambda: None
    app_paths.get_user_data_dir = lambda: user_data_dir
    sys.modules["app_paths"] = app_paths


def _import_crop():
    os.environ["API_KEY"] = "test-key"
    sys.modules.pop("crop", None)
    _install_crop_import_fakes()
    return importlib.import_module("crop")


def _import_scan2():
    sys.modules.pop("scan2", None)
    _install_scan2_import_fakes()
    return importlib.import_module("scan2")


def _install_barcode_import_fakes():
    cv2 = types.ModuleType("cv2")
    cv2.COLOR_BGR2GRAY = 1
    cv2.COLOR_GRAY2BGR = 2
    cv2.BORDER_CONSTANT = 0
    cv2.INTER_CUBIC = 0
    cv2.ADAPTIVE_THRESH_GAUSSIAN_C = 0
    cv2.THRESH_BINARY = 0
    cv2.MORPH_RECT = 0
    cv2.MORPH_CLOSE = 0
    cv2.ROTATE_90_CLOCKWISE = 0
    cv2.cvtColor = lambda img, *args, **kwargs: img
    cv2.resize = lambda img, *args, **kwargs: img
    cv2.bitwise_not = lambda img: img
    cv2.copyMakeBorder = lambda img, *args, **kwargs: img
    cv2.createCLAHE = lambda *args, **kwargs: types.SimpleNamespace(apply=lambda img: img)
    cv2.GaussianBlur = lambda img, *args, **kwargs: img
    cv2.addWeighted = lambda img, *args, **kwargs: img
    cv2.adaptiveThreshold = lambda img, *args, **kwargs: img
    cv2.getStructuringElement = lambda *args, **kwargs: object()
    cv2.morphologyEx = lambda img, *args, **kwargs: img
    cv2.imwrite = lambda *args, **kwargs: True
    sys.modules["cv2"] = cv2

    numpy = types.ModuleType("numpy")
    numpy.ndarray = object
    numpy.rot90 = lambda img, k: img
    sys.modules["numpy"] = numpy

    pyzbar_pkg = types.ModuleType("pyzbar")
    pyzbar_mod = types.ModuleType("pyzbar.pyzbar")
    pyzbar_mod.ZBarSymbol = types.SimpleNamespace(CODE128=object())
    pyzbar_mod.decode = lambda *args, **kwargs: []
    pyzbar_pkg.pyzbar = pyzbar_mod
    sys.modules["pyzbar"] = pyzbar_pkg
    sys.modules["pyzbar.pyzbar"] = pyzbar_mod


def _import_barcode():
    sys.modules.pop("barcode", None)
    sys.modules.pop("huaweiocr.barcode.generic", None)
    _install_barcode_import_fakes()
    return importlib.import_module("barcode")


@unittest.skipUnless(os.name == "nt", "Windows file locking behavior only")
class LockedOutputDirsTest(unittest.TestCase):
    def test_locked_manifest_switches_stage2_to_run_directory(self):
        crop = _import_crop()
        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "new_images")
            original_stage2 = os.path.join(root, "stage2_fields")
            os.makedirs(original_stage2)
            locked_manifest = os.path.join(original_stage2, "manifest.jsonl")

            with open(locked_manifest, "w", encoding="utf-8") as locked:
                locked.write("locked\n")
                locked.flush()

                crop.configure_paths(input_dir=input_dir, out_dir=root)
                crop.ensure_dirs()

                self.assertNotEqual(
                    os.path.abspath(original_stage2),
                    os.path.abspath(crop.STAGE2_DIR),
                )
                self.assertTrue(os.path.isdir(crop.OUT_MODEL_DIR))
                self.assertTrue(os.path.isdir(crop.OUT_SN_DIR))
                self.assertTrue(os.path.isdir(crop.FAILED_DIR))
                self.assertEqual(
                    crop.MANIFEST_PATH,
                    os.path.join(crop.STAGE2_DIR, "manifest.jsonl"),
                )
                self.assertTrue(os.path.exists(locked_manifest))


class RunAllPathPropagationTest(unittest.TestCase):
    def test_scan2_reads_actual_crop_output_dirs(self):
        import run_all

        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "input")
            out_dir = os.path.join(root, "out")
            os.makedirs(input_dir)
            with open(os.path.join(input_dir, "sample.png"), "wb") as image:
                image.write(b"not-a-real-image")

            crop = types.ModuleType("crop")
            crop.OUT_MODEL_DIR = os.path.join(out_dir, "stage2_fields_run_x", "model")
            crop.OUT_SN_DIR = os.path.join(out_dir, "stage2_fields_run_x", "sn")
            crop.STAGE2_DIR = os.path.join(out_dir, "stage2_fields_run_x")
            crop.STAGE1_DIR = os.path.join(out_dir, "stage1_labels_run_x")
            crop.set_log_level = lambda level: None
            crop.main = lambda **kwargs: {
                "label_count": 1,
                "manifest_rows": 1,
                "stage1_dir": crop.STAGE1_DIR,
                "stage2_dir": crop.STAGE2_DIR,
                "model_dir": crop.OUT_MODEL_DIR,
                "sn_dir": crop.OUT_SN_DIR,
            }

            calls = {}
            scan2 = types.ModuleType("scan2")
            scan2.set_log_level = lambda level: None

            def fake_scan2_main(**kwargs):
                calls.update(kwargs)
                return {"sn_total": 0}

            scan2.main = fake_scan2_main

            argv = [
                "run_all.py",
                "--input",
                input_dir,
                "--out",
                out_dir,
            ]
            with mock.patch.dict(sys.modules, {"crop": crop, "scan2": scan2}):
                with mock.patch.object(sys, "argv", argv):
                    self.assertEqual(run_all.main(), 0)

            summary_path = os.path.join(out_dir, "run_summary.json")
            with open(summary_path, "r", encoding="utf-8") as f:
                summary = json.load(f)

            self.assertEqual(calls["model_dir"], crop.OUT_MODEL_DIR)
            self.assertEqual(calls["sn_dir"], crop.OUT_SN_DIR)
            self.assertEqual(
                calls["out_jsonl"],
                os.path.join(crop.STAGE2_DIR, "model_sn_ocr.jsonl"),
            )
            self.assertEqual(summary["exit_status"], 0)
            self.assertEqual(summary["status"], "success")
            self.assertEqual(summary["image_count"], 1)
            self.assertEqual(summary["crop_stats"]["label_count"], 1)
            self.assertEqual(summary["scan2_stats"]["sn_total"], 0)
            self.assertEqual(summary["output_paths"]["model_dir"], crop.OUT_MODEL_DIR)
            self.assertEqual(summary["output_paths"]["result_jsonl"], calls["out_jsonl"])
            self.assertIsInstance(summary["timing_sec"]["crop"], float)
            self.assertIsInstance(summary["timing_sec"]["scan"], float)


class Scan2MapOrderedProgressTest(unittest.TestCase):
    def test_map_ordered_reports_progress_and_preserves_input_order(self):
        scan2 = _import_scan2()

        def work(item):
            time.sleep({1: 0.03, 2: 0.01, 3: 0.02}[item])
            return item * 10

        progress = []
        results = scan2._map_ordered(
            [1, 2, 3],
            work,
            workers=3,
            progress=lambda done, total: progress.append((done, total)),
        )

        self.assertEqual(results, [10, 20, 30])
        self.assertEqual(progress, [(1, 3), (2, 3), (3, 3)])

    def test_zero_label_crop_returns_nonzero_without_scanning(self):
        import run_all

        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "input")
            out_dir = os.path.join(root, "out")
            os.makedirs(input_dir)
            with open(os.path.join(input_dir, "sample.png"), "wb") as image:
                image.write(b"not-a-real-image")

            crop = types.ModuleType("crop")
            crop.set_log_level = lambda level: None
            crop.main = mock.Mock(return_value={"label_count": 0, "manifest_rows": 0})

            scan2 = types.ModuleType("scan2")
            scan2.set_log_level = lambda level: None
            scan2.main = mock.Mock()

            argv = ["run_all.py", "--input", input_dir, "--out", out_dir]
            with mock.patch.dict(sys.modules, {"crop": crop, "scan2": scan2}):
                with mock.patch.object(sys, "argv", argv):
                    self.assertEqual(run_all.main(), 1)

            with open(os.path.join(out_dir, "run_summary.json"), "r", encoding="utf-8") as f:
                summary = json.load(f)

            scan2.main.assert_not_called()
            self.assertEqual(summary["exit_status"], 1)
            self.assertEqual(summary["status"], "failed")
            self.assertEqual(summary["crop_stats"]["label_count"], 0)
            self.assertIn("No label crops", summary["message"])

    def test_summary_json_out_overrides_default_path(self):
        import run_all

        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "missing")
            out_dir = os.path.join(root, "out")
            summary_path = os.path.join(root, "agent", "summary.json")

            argv = [
                "run_all.py",
                "--input",
                input_dir,
                "--out",
                out_dir,
                "--summary-json-out",
                summary_path,
            ]
            with mock.patch.object(sys, "argv", argv):
                self.assertEqual(run_all.main(), 2)

            with open(summary_path, "r", encoding="utf-8") as f:
                summary = json.load(f)
            self.assertEqual(summary["exit_status"], 2)
            self.assertEqual(summary["output_paths"]["summary_json"], summary_path)
            self.assertFalse(os.path.exists(os.path.join(out_dir, "run_summary.json")))

    def test_successful_pipeline_returns_nonzero_when_summary_path_is_directory(self):
        import run_all

        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "input")
            out_dir = os.path.join(root, "out")
            summary_path = os.path.join(root, "summary_dir")
            os.makedirs(input_dir)
            os.makedirs(summary_path)
            with open(os.path.join(input_dir, "sample.png"), "wb") as image:
                image.write(b"not-a-real-image")

            crop = types.ModuleType("crop")
            crop.OUT_MODEL_DIR = os.path.join(out_dir, "stage2_fields", "model")
            crop.OUT_SN_DIR = os.path.join(out_dir, "stage2_fields", "sn")
            crop.STAGE2_DIR = os.path.join(out_dir, "stage2_fields")
            crop.STAGE1_DIR = os.path.join(out_dir, "stage1_labels")
            crop.set_log_level = lambda level: None
            crop.main = lambda **kwargs: {
                "label_count": 1,
                "manifest_rows": 1,
                "stage1_dir": crop.STAGE1_DIR,
                "stage2_dir": crop.STAGE2_DIR,
                "model_dir": crop.OUT_MODEL_DIR,
                "sn_dir": crop.OUT_SN_DIR,
            }

            scan2 = types.ModuleType("scan2")
            scan2.set_log_level = lambda level: None
            scan2.main = lambda **kwargs: {"sn_total": 0}

            argv = [
                "run_all.py",
                "--input",
                input_dir,
                "--out",
                out_dir,
                "--summary-json-out",
                summary_path,
            ]
            stderr = io.StringIO()
            with mock.patch.dict(sys.modules, {"crop": crop, "scan2": scan2}):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch("sys.stderr", stderr):
                        self.assertEqual(run_all.main(), 1)

            stderr_text = stderr.getvalue()
            self.assertIn("Warning: failed to write run summary", stderr_text)
            self.assertIn("Error: failed to write run summary", stderr_text)

    def test_existing_failure_keeps_exit_code_when_summary_path_is_directory(self):
        import run_all

        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "missing")
            out_dir = os.path.join(root, "out")
            summary_path = os.path.join(root, "summary_dir")
            os.makedirs(summary_path)

            argv = [
                "run_all.py",
                "--input",
                input_dir,
                "--out",
                out_dir,
                "--summary-json-out",
                summary_path,
            ]
            stderr = io.StringIO()
            with mock.patch.object(sys, "argv", argv):
                with mock.patch("sys.stderr", stderr):
                    self.assertEqual(run_all.main(), 2)

            stderr_text = stderr.getvalue()
            self.assertIn("Input directory does not exist", stderr_text)
            self.assertIn("Warning: failed to write run summary", stderr_text)
            self.assertNotIn("returning failure status", stderr_text)

    def test_cli_can_export_gui_equivalent_excel(self):
        import run_all

        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "input")
            out_dir = os.path.join(root, "out")
            os.makedirs(input_dir)
            with open(os.path.join(input_dir, "sample.png"), "wb") as image:
                image.write(b"not-a-real-image")

            crop = types.ModuleType("crop")
            crop.OUT_MODEL_DIR = os.path.join(out_dir, "stage2_fields", "model")
            crop.OUT_SN_DIR = os.path.join(out_dir, "stage2_fields", "sn")
            crop.STAGE2_DIR = os.path.join(out_dir, "stage2_fields")
            crop.STAGE1_DIR = os.path.join(out_dir, "stage1_labels")
            crop.set_log_level = lambda level: None
            crop.main = lambda **kwargs: {
                "label_count": 1,
                "manifest_rows": 1,
                "stage1_dir": crop.STAGE1_DIR,
                "stage2_dir": crop.STAGE2_DIR,
                "model_dir": crop.OUT_MODEL_DIR,
                "sn_dir": crop.OUT_SN_DIR,
            }

            scan2 = types.ModuleType("scan2")
            scan2.set_log_level = lambda level: None

            def fake_scan2_main(**kwargs):
                os.makedirs(os.path.dirname(kwargs["out_jsonl"]), exist_ok=True)
                with open(kwargs["out_jsonl"], "w", encoding="utf-8") as f:
                    f.write(json.dumps({
                        "label_id": "sample.png__label_1",
                        "model": "AP162E",
                        "sn": "4E25A0170000",
                        "model_src": "ocr_file",
                        "sn_src": "barcode",
                    }) + "\n")
                return {"records": 1, "sn_total": 1, "sn_success": 1}

            scan2.main = fake_scan2_main

            excel_path = os.path.join(root, "result.xlsx")
            argv = [
                "run_all.py",
                "--input",
                input_dir,
                "--out",
                out_dir,
                "--excel-out",
                excel_path,
            ]
            with mock.patch.dict(sys.modules, {"crop": crop, "scan2": scan2}):
                with mock.patch.object(sys, "argv", argv):
                    self.assertEqual(run_all.main(), 0)

            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            self.assertEqual(
                [cell.value for cell in ws[1]],
                ["label_id", "model", "sn", "model_src", "sn_src"],
            )
            self.assertEqual(
                [cell.value for cell in ws[2]],
                ["sample.png__label_1", "AP162E", "4E25A0170000", "ocr_file", "barcode"],
            )

            with open(os.path.join(out_dir, "run_summary.json"), "r", encoding="utf-8") as f:
                summary = json.load(f)
            self.assertEqual(summary["excel_export"]["rows"], 1)
            self.assertEqual(summary["output_paths"]["excel"], excel_path)

    def test_empty_input_returns_nonzero_without_running_pipeline(self):
        import run_all

        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "empty")
            os.makedirs(input_dir)
            crop = types.ModuleType("crop")
            crop.set_log_level = lambda level: None
            crop.main = mock.Mock()

            scan2 = types.ModuleType("scan2")
            scan2.set_log_level = lambda level: None
            scan2.main = mock.Mock()

            argv = ["run_all.py", "--input", input_dir, "--out", os.path.join(root, "out")]
            with mock.patch.dict(sys.modules, {"crop": crop, "scan2": scan2}):
                with mock.patch.object(sys, "argv", argv):
                    self.assertEqual(run_all.main(), 2)

            crop.main.assert_not_called()
            scan2.main.assert_not_called()


class PaddleOcrModelKwargsTest(unittest.TestCase):
    def test_torch_compat_stub_exposes_import_spec(self):
        for name in (
            "ocr",
            "huaweiocr.detect.ocr_engine",
            "paddle",
            "paddleocr",
            "app_paths",
            "huaweiocr.io.paths_runtime",
            "torch",
            "torch.multiprocessing",
            "torch.distributed",
            "torch.nn",
            "torch.nn.functional",
            "torch.nn.utils",
            "torch.nn.utils.rnn",
        ):
            sys.modules.pop(name, None)

        torch = types.ModuleType("torch")
        sys.modules["torch"] = torch

        paddle = types.ModuleType("paddle")
        paddle.set_device = lambda _device: None
        sys.modules["paddle"] = paddle

        paddleocr = types.ModuleType("paddleocr")
        paddleocr.PaddleOCR = type("DummyPaddleOCR", (), {})
        sys.modules["paddleocr"] = paddleocr

        app_paths = types.ModuleType("app_paths")
        app_paths.ensure_models_installed = lambda: None
        app_paths.get_resource_path = lambda *parts: os.path.join(*parts)
        sys.modules["app_paths"] = app_paths

        import ocr  # noqa: F401  (import succeeding under stubs is the assertion)

        torch_stub = sys.modules["torch"]
        self.assertIsNotNone(torch_stub.__spec__)
        self.assertEqual(torch_stub.__spec__.name, "torch")
        self.assertTrue(torch_stub.__spec__.submodule_search_locations is not None)
        self.assertEqual(importlib.util.find_spec("torch").name, "torch")
        self.assertEqual(importlib.util.find_spec("torch.multiprocessing").name, "torch.multiprocessing")
        self.assertEqual(importlib.util.find_spec("torch.distributed").name, "torch.distributed")
        self.assertEqual(importlib.util.find_spec("torch.nn").name, "torch.nn")
        self.assertEqual(importlib.util.find_spec("torch.nn.functional").name, "torch.nn.functional")
        self.assertEqual(importlib.util.find_spec("torch.nn.utils.rnn").name, "torch.nn.utils.rnn")
        self.assertIs(torch_stub.nn.Module, torch_stub.nn.Linear)
        self.assertIs(torch_stub.Tensor, torch_stub.ByteTensor)
        self.assertIsInstance(torch_stub.tensor(), torch_stub.Tensor)
        self.assertTrue(callable(torch_stub.nn.utils.rnn.pad_sequence))
        self.assertTrue(callable(torch_stub.no_grad))

    def test_torch_compat_stub_allows_modelscope_torch_utils_import(self):
        if importlib.util.find_spec("modelscope.utils.torch_utils") is None:
            self.skipTest("modelscope is not installed")

        for name in list(sys.modules):
            if name in ("ocr", "huaweiocr.detect.ocr_engine", "paddle", "paddleocr", "app_paths", "huaweiocr.io.paths_runtime") or name.startswith("torch") or name.startswith("modelscope"):
                sys.modules.pop(name, None)

        sys.modules["torch"] = types.ModuleType("torch")

        paddle = types.ModuleType("paddle")
        paddle.set_device = lambda _device: None
        sys.modules["paddle"] = paddle

        paddleocr = types.ModuleType("paddleocr")
        paddleocr.PaddleOCR = type("DummyPaddleOCR", (), {})
        sys.modules["paddleocr"] = paddleocr

        app_paths = types.ModuleType("app_paths")
        app_paths.ensure_models_installed = lambda: None
        app_paths.get_resource_path = lambda *parts: os.path.join(*parts)
        sys.modules["app_paths"] = app_paths

        import ocr  # noqa: F401  (import succeeding under stubs is the assertion)

        torch_utils = importlib.import_module("modelscope.utils.torch_utils")
        self.assertTrue(hasattr(torch_utils, "is_on_same_device"))
        self.assertIs(sys.modules["torch"].nn.Module, sys.modules["torch"].nn.Linear)

    def test_local_model_dirs_include_matching_model_names(self):
        for name in ("ocr", "huaweiocr.detect.ocr_engine", "paddle", "paddleocr", "app_paths", "huaweiocr.io.paths_runtime"):
            sys.modules.pop(name, None)

        paddle = types.ModuleType("paddle")
        paddle.set_device = lambda _device: None
        sys.modules["paddle"] = paddle

        paddleocr = types.ModuleType("paddleocr")

        class DummyPaddleOCR:
            def __init__(
                self,
                use_doc_orientation_classify=None,
                use_doc_unwarping=None,
                use_textline_orientation=None,
                text_detection_model_name=None,
                text_detection_model_dir=None,
                text_recognition_model_name=None,
                text_recognition_model_dir=None,
                textline_orientation_model_name=None,
                textline_orientation_model_dir=None,
            ):
                pass

        paddleocr.PaddleOCR = DummyPaddleOCR
        sys.modules["paddleocr"] = paddleocr

        app_paths = types.ModuleType("app_paths")
        app_paths.ensure_models_installed = lambda: None
        app_paths.get_resource_path = lambda *parts: os.path.join(*parts)
        sys.modules["app_paths"] = app_paths

        import ocr  # noqa: F401  (import succeeding under stubs is the assertion)

        with tempfile.TemporaryDirectory() as root:
            models = os.path.join(root, "official_models")
            for name in (
                "PP-OCRv5_server_det",
                "PP-OCRv5_server_rec",
                "en_PP-OCRv5_mobile_rec",
                "PP-LCNet_x1_0_textline_ori",
            ):
                os.makedirs(os.path.join(models, name))

            kwargs = ocr._paddleocr_model_kwargs(models)

        self.assertEqual(kwargs["text_detection_model_name"], "PP-OCRv5_server_det")
        self.assertEqual(kwargs["text_recognition_model_name"], "en_PP-OCRv5_mobile_rec")
        self.assertEqual(kwargs["textline_orientation_model_name"], "PP-LCNet_x1_0_textline_ori")
        self.assertIs(kwargs["use_doc_orientation_classify"], False)
        self.assertIs(kwargs["use_doc_unwarping"], False)
        self.assertIs(kwargs["use_textline_orientation"], True)
        self.assertTrue(kwargs["text_detection_model_dir"].endswith("PP-OCRv5_server_det"))
        self.assertTrue(kwargs["text_recognition_model_dir"].endswith("en_PP-OCRv5_mobile_rec"))
        self.assertTrue(kwargs["textline_orientation_model_dir"].endswith("PP-LCNet_x1_0_textline_ori"))

    def test_fast_profile_prefers_mobile_recognition_model(self):
        for name in ("ocr", "huaweiocr.detect.ocr_engine", "paddle", "paddleocr", "app_paths", "huaweiocr.io.paths_runtime"):
            sys.modules.pop(name, None)

        paddle = types.ModuleType("paddle")
        paddle.set_device = lambda _device: None
        sys.modules["paddle"] = paddle

        paddleocr = types.ModuleType("paddleocr")

        class DummyPaddleOCR:
            def __init__(self, text_recognition_model_name=None, text_recognition_model_dir=None):
                pass

        paddleocr.PaddleOCR = DummyPaddleOCR
        sys.modules["paddleocr"] = paddleocr

        app_paths = types.ModuleType("app_paths")
        app_paths.ensure_models_installed = lambda: None
        app_paths.get_resource_path = lambda *parts: os.path.join(*parts)
        sys.modules["app_paths"] = app_paths

        with mock.patch.dict(os.environ, {"HUAWEIOCR_OCR_PROFILE": "fast"}, clear=False):
            import ocr  # noqa: F401  (import succeeding under stubs is the assertion)

        with tempfile.TemporaryDirectory() as root:
            models = os.path.join(root, "official_models")
            for name in ("PP-OCRv5_server_rec", "en_PP-OCRv5_mobile_rec"):
                os.makedirs(os.path.join(models, name))

            kwargs = ocr._paddleocr_model_kwargs(models)

        self.assertEqual(kwargs["text_recognition_model_name"], "en_PP-OCRv5_mobile_rec")
        self.assertTrue(kwargs["text_recognition_model_dir"].endswith("en_PP-OCRv5_mobile_rec"))

    def test_server_profile_prefers_server_recognition_model(self):
        for name in ("ocr", "huaweiocr.detect.ocr_engine", "paddle", "paddleocr", "app_paths", "huaweiocr.io.paths_runtime"):
            sys.modules.pop(name, None)

        paddle = types.ModuleType("paddle")
        paddle.set_device = lambda _device: None
        sys.modules["paddle"] = paddle

        paddleocr = types.ModuleType("paddleocr")

        class DummyPaddleOCR:
            def __init__(self, text_recognition_model_name=None, text_recognition_model_dir=None):
                pass

        paddleocr.PaddleOCR = DummyPaddleOCR
        sys.modules["paddleocr"] = paddleocr

        app_paths = types.ModuleType("app_paths")
        app_paths.ensure_models_installed = lambda: None
        app_paths.get_resource_path = lambda *parts: os.path.join(*parts)
        sys.modules["app_paths"] = app_paths

        with mock.patch.dict(os.environ, {"HUAWEIOCR_OCR_PROFILE": "server"}, clear=False):
            import ocr  # noqa: F401  (import succeeding under stubs is the assertion)

        with tempfile.TemporaryDirectory() as root:
            models = os.path.join(root, "official_models")
            for name in ("PP-OCRv5_server_rec", "en_PP-OCRv5_mobile_rec"):
                os.makedirs(os.path.join(models, name))

            kwargs = ocr._paddleocr_model_kwargs(models)

        self.assertEqual(kwargs["text_recognition_model_name"], "PP-OCRv5_server_rec")
        self.assertTrue(kwargs["text_recognition_model_dir"].endswith("PP-OCRv5_server_rec"))

    def test_local_source_model_root_falls_back_to_bundle_models(self):
        for name in ("ocr", "huaweiocr.detect.ocr_engine", "paddle", "paddleocr", "app_paths", "huaweiocr.io.paths_runtime"):
            sys.modules.pop(name, None)

        paddle = types.ModuleType("paddle")
        paddle.set_device = lambda _device: None
        sys.modules["paddle"] = paddle

        paddleocr = types.ModuleType("paddleocr")
        paddleocr.PaddleOCR = type("DummyPaddleOCR", (), {})
        sys.modules["paddleocr"] = paddleocr

        app_paths = types.ModuleType("app_paths")
        app_paths.ensure_models_installed = lambda: None
        app_paths.get_resource_path = lambda *parts: os.path.join(*parts)
        sys.modules["app_paths"] = app_paths

        import ocr  # noqa: F401  (import succeeding under stubs is the assertion)

        with tempfile.TemporaryDirectory() as root:
            bundle_root = os.path.join(root, "bundle", "models", "official_models")
            os.makedirs(bundle_root)
            with mock.patch.object(ocr, "get_resource_path", side_effect=lambda *parts: os.path.join(root, *parts)):
                self.assertEqual(ocr._local_model_root_fallback(), bundle_root)

    def test_ocr_import_disables_packaged_onednn_executor_path(self):
        for name in ("ocr", "huaweiocr.detect.ocr_engine", "paddle", "paddleocr", "app_paths", "huaweiocr.io.paths_runtime"):
            sys.modules.pop(name, None)

        for name in (
            "FLAGS_use_mkldnn",
            "FLAGS_use_onednn",
            "FLAGS_enable_pir_api",
            "FLAGS_enable_pir_in_executor",
            "PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT",
        ):
            os.environ.pop(name, None)

        paddle = types.ModuleType("paddle")
        paddle.set_device = lambda _device: None
        sys.modules["paddle"] = paddle

        paddleocr = types.ModuleType("paddleocr")
        paddleocr.PaddleOCR = type("DummyPaddleOCR", (), {})
        sys.modules["paddleocr"] = paddleocr

        app_paths = types.ModuleType("app_paths")
        app_paths.ensure_models_installed = lambda: None
        app_paths.get_resource_path = lambda *parts: os.path.join(*parts)
        sys.modules["app_paths"] = app_paths

        import ocr  # noqa: F401  (import succeeding under stubs is the assertion)

        self.assertEqual(os.environ["FLAGS_use_mkldnn"], "0")
        self.assertEqual(os.environ["FLAGS_use_onednn"], "0")
        self.assertEqual(os.environ["FLAGS_enable_pir_api"], "0")
        self.assertEqual(os.environ["FLAGS_enable_pir_in_executor"], "0")
        self.assertEqual(os.environ["PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT"], "False")


class Scan2ManifestTest(unittest.TestCase):
    def test_main_signature_keeps_legacy_arguments(self):
        scan2 = _import_scan2()

        signature = inspect.signature(scan2.main)
        self.assertEqual(
            list(signature.parameters),
            ["out_dir", "model_dir", "sn_dir", "out_jsonl", "debug_log", "log_level"],
        )
        self.assertEqual(
            str(signature),
            "(out_dir=None, model_dir=None, sn_dir=None, out_jsonl=None, debug_log=None, log_level='info')",
        )

    def test_read_image_prefers_unicode_safe_path_for_non_ascii_filename(self):
        scan2 = _import_scan2()
        path = r"F:\HuaweiOCR\stage2_fields\sn\滑坳掳.jpg__label_1__sn.png"
        sentinel = object()
        scan2.np.uint8 = object()

        with mock.patch.object(scan2.np, "fromfile", return_value=b"123", create=True) as fromfile:
            with mock.patch.object(scan2.cv2, "imdecode", return_value=sentinel, create=True) as imdecode:
                with mock.patch.object(
                    scan2.cv2,
                    "imread",
                    side_effect=AssertionError("non-ascii path should bypass cv2.imread first"),
                    create=True,
                ):
                    self.assertIs(scan2._read_image(path), sentinel)

        fromfile.assert_called_once_with(path, dtype=scan2.np.uint8)
        imdecode.assert_called_once()

    def test_manifest_keeps_labels_without_model_or_sn_crops(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)

            manifest_path = os.path.join(stage2, "manifest.jsonl")
            with open(manifest_path, "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": "img__with__sep__label_1"}) + "\n")
                manifest.write(json.dumps({"label_id": "missing_both__label_1"}) + "\n")

            with open(os.path.join(model_dir, "img__with__sep__label_1__model.png"), "wb") as image:
                image.write(b"model")

            out_jsonl = os.path.join(root, "out.jsonl")
            with mock.patch.object(scan2, "recognize_model_ocr", return_value=("M1", "raw", "test")):
                with mock.patch.dict(os.environ, {"SCAN2_MODEL_BARCODE": "0"}, clear=False):
                    stats = scan2.main(
                        model_dir=model_dir,
                        sn_dir=sn_dir,
                        out_jsonl=out_jsonl,
                        debug_log=os.path.join(root, "debug.log"),
                    )

            with open(out_jsonl, "r", encoding="utf-8") as f:
                rows = [json.loads(line) for line in f if line.strip()]

            self.assertEqual([row["label_id"] for row in rows], ["img__with__sep__label_1", "missing_both__label_1"])
            self.assertEqual(rows[1]["model_src"], "missing")
            self.assertEqual(rows[1]["sn_src"], "missing")
            self.assertEqual(stats["sn_total"], 0)

    def test_manifest_bad_json_fails_fast(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            os.makedirs(os.path.join(stage2, "model"))
            os.makedirs(os.path.join(stage2, "sn"))
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write("{bad json}\n")

            with self.assertRaisesRegex(ValueError, "Invalid manifest JSON"):
                scan2.main(
                    model_dir=os.path.join(stage2, "model"),
                    sn_dir=os.path.join(stage2, "sn"),
                    out_jsonl=os.path.join(root, "out.jsonl"),
                    debug_log=os.path.join(root, "debug.log"),
                )

    def test_manifest_missing_label_id_fails_fast(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            os.makedirs(os.path.join(stage2, "model"))
            os.makedirs(os.path.join(stage2, "sn"))
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"model_path": None}) + "\n")

            with self.assertRaisesRegex(ValueError, "missing label_id"):
                scan2.main(
                    model_dir=os.path.join(stage2, "model"),
                    sn_dir=os.path.join(stage2, "sn"),
                    out_jsonl=os.path.join(root, "out.jsonl"),
                    debug_log=os.path.join(root, "debug.log"),
                )

    def test_raw_fields_are_masked_by_default(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            model_path = os.path.join(model_dir, "a__label_1__model.png")
            sn_path = os.path.join(sn_dir, "a__label_1__sn.png")
            open(model_path, "wb").close()
            open(sn_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": "a__label_1", "model_path": model_path, "sn_path": sn_path}) + "\n")

            out_jsonl = os.path.join(root, "out.jsonl")
            meta = {"barcode_found": False, "ocr_text_found": False, "barcode_status": "decoder_miss"}
            report = types.SimpleNamespace(status="decoder_miss", results=[])
            with mock.patch.object(scan2, "recognize_model_ocr", return_value=("AP162E", "RAW_MODEL_SECRET_123456", "ocr_file")):
                with mock.patch.object(scan2, "_recognize_sn_barcode", return_value=("", "", "barcode_decoder_miss", meta, report)):
                    with mock.patch.object(scan2, "recognize_sn_ocr_after_barcode", return_value=("4E25A0170000", "RAW_SN_SECRET_123456", "ocr", meta)):
                        with mock.patch.dict(
                            os.environ,
                            {
                                "SCAN2_MASK_RAW": "",
                                "SCAN2_UNSAFE_RAW": "",
                                "HUAWEIOCR_UNSAFE_RAW": "",
                                "SCAN2_MODEL_BARCODE": "0",
                                "SCAN2_OCR_FALLBACK": "1",
                            },
                            clear=False,
                        ):
                            scan2.main(
                                model_dir=model_dir,
                                sn_dir=sn_dir,
                                out_jsonl=out_jsonl,
                                debug_log=os.path.join(root, "debug.log"),
                            )

            with open(out_jsonl, "r", encoding="utf-8") as f:
                row = json.loads(f.readline())
            self.assertEqual(row["model"], "AP162E")
            self.assertEqual(row["sn"], "4E25A0170000")
            self.assertNotIn("RAW_MODEL_SECRET_123456", row["model_raw"])
            self.assertNotIn("RAW_SN_SECRET_123456", row["sn_raw"])

    def test_raw_fields_can_keep_full_values_with_unsafe_env(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            model_path = os.path.join(model_dir, "a__label_1__model.png")
            sn_path = os.path.join(sn_dir, "a__label_1__sn.png")
            open(model_path, "wb").close()
            open(sn_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": "a__label_1", "model_path": model_path, "sn_path": sn_path}) + "\n")

            out_jsonl = os.path.join(root, "out.jsonl")
            meta = {"barcode_found": False, "ocr_text_found": False, "barcode_status": "decoder_miss"}
            report = types.SimpleNamespace(status="decoder_miss", results=[])
            with mock.patch.object(scan2, "recognize_model_ocr", return_value=("AP162E", "RAW_MODEL_SECRET_123456", "ocr_file")):
                with mock.patch.object(scan2, "_recognize_sn_barcode", return_value=("", "", "barcode_decoder_miss", meta, report)):
                    with mock.patch.object(scan2, "recognize_sn_ocr_after_barcode", return_value=("4E25A0170000", "RAW_SN_SECRET_123456", "ocr", meta)):
                        with mock.patch.dict(
                            os.environ,
                            {
                                "SCAN2_MASK_RAW": "",
                                "SCAN2_UNSAFE_RAW": "1",
                                "HUAWEIOCR_UNSAFE_RAW": "",
                                "SCAN2_MODEL_BARCODE": "0",
                                "SCAN2_OCR_FALLBACK": "1",
                            },
                            clear=False,
                        ):
                            scan2.main(
                                model_dir=model_dir,
                                sn_dir=sn_dir,
                                out_jsonl=out_jsonl,
                                debug_log=os.path.join(root, "debug.log"),
                            )

            with open(out_jsonl, "r", encoding="utf-8") as f:
                row = json.loads(f.readline())
            self.assertEqual(row["model_raw"], "RAW_MODEL_SECRET_123456")
            self.assertEqual(row["sn_raw"], "RAW_SN_SECRET_123456")

    def test_huaweiocr_unsafe_raw_alias_disables_raw_field_masking(self):
        scan2 = _import_scan2()

        with mock.patch.dict(
            os.environ,
            {
                "SCAN2_UNSAFE_RAW": "",
                "HUAWEIOCR_UNSAFE_RAW": "1",
            },
            clear=False,
        ):
            self.assertFalse(scan2.raw_result_fields_are_masked())

    def test_raw_fields_can_be_masked_with_env_flag(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            model_path = os.path.join(model_dir, "a__label_1__model.png")
            sn_path = os.path.join(sn_dir, "a__label_1__sn.png")
            open(model_path, "wb").close()
            open(sn_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": "a__label_1", "model_path": model_path, "sn_path": sn_path}) + "\n")

            out_jsonl = os.path.join(root, "out.jsonl")
            meta = {"barcode_found": False, "ocr_text_found": False, "barcode_status": "decoder_miss"}
            report = types.SimpleNamespace(status="decoder_miss", results=[])
            with mock.patch.object(scan2, "recognize_model_ocr", return_value=("AP162E", "RAW_MODEL_SECRET_123456", "ocr_file")):
                with mock.patch.object(scan2, "_recognize_sn_barcode", return_value=("", "", "barcode_decoder_miss", meta, report)):
                    with mock.patch.object(scan2, "recognize_sn_ocr_after_barcode", return_value=("4E25A0170000", "RAW_SN_SECRET_123456", "ocr", meta)):
                        with mock.patch.dict(
                            os.environ,
                            {
                                "SCAN2_MASK_RAW": "1",
                                "SCAN2_UNSAFE_RAW": "",
                                "HUAWEIOCR_UNSAFE_RAW": "",
                                "SCAN2_MODEL_BARCODE": "0",
                                "SCAN2_OCR_FALLBACK": "1",
                            },
                            clear=False,
                        ):
                            scan2.main(
                                model_dir=model_dir,
                                sn_dir=sn_dir,
                                out_jsonl=out_jsonl,
                                debug_log=os.path.join(root, "debug.log"),
                            )

            with open(out_jsonl, "r", encoding="utf-8") as f:
                row = json.loads(f.readline())
            self.assertNotIn("RAW_MODEL_SECRET_123456", row["model_raw"])
            self.assertNotIn("RAW_SN_SECRET_123456", row["sn_raw"])

    def test_info_log_keeps_full_model_and_sn_values(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            model_path = os.path.join(model_dir, "a__label_1__model.png")
            sn_path = os.path.join(sn_dir, "a__label_1__sn.png")
            open(model_path, "wb").close()
            open(sn_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": "a__label_1", "model_path": model_path, "sn_path": sn_path}) + "\n")

            logs = []
            old_sink = scan2.LOG_SINK
            scan2.set_log_sink(logs.append)
            try:
                meta = {"barcode_found": True, "ocr_text_found": False, "barcode_status": "hit"}
                report = types.SimpleNamespace(status="hit", results=[])
                with mock.patch.object(scan2, "recognize_model_ocr", return_value=("S380-S8P2T", "raw", "ocr_color")):
                    with mock.patch.object(scan2, "_recognize_sn_barcode", return_value=("4E25B0105849", "raw", "barcode", meta, report)):
                        with mock.patch.dict(os.environ, {"SCAN2_MODEL_BARCODE": "0"}, clear=False):
                            scan2.main(
                                model_dir=model_dir,
                                sn_dir=sn_dir,
                                out_jsonl=os.path.join(root, "out.jsonl"),
                                debug_log=os.path.join(root, "debug.log"),
                            )
            finally:
                scan2.set_log_sink(old_sink)

            joined = "\n".join(logs)
            self.assertIn("型号=S380-S8P2T", joined)
            self.assertIn("SN=4E25B0105849", joined)
            self.assertNotIn("S380**8P2T", joined)
            self.assertNotIn("4E25****5849", joined)

    def test_info_log_emits_realtime_barcode_progress(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            model_path = os.path.join(model_dir, "a__label_1__model.png")
            sn_path = os.path.join(sn_dir, "a__label_1__sn.png")
            open(model_path, "wb").close()
            open(sn_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": "a__label_1", "model_path": model_path, "sn_path": sn_path}) + "\n")

            logs = []
            old_sink = scan2.LOG_SINK
            scan2.set_log_sink(logs.append)
            try:
                meta = {"barcode_found": True, "ocr_text_found": False, "barcode_status": "hit"}
                report = types.SimpleNamespace(status="hit", results=[])
                with mock.patch.object(scan2, "recognize_model_ocr", return_value=("S380-S8P2T", "raw", "ocr_color")):
                    with mock.patch.object(scan2, "_recognize_sn_barcode", return_value=("4E25B0105849", "raw", "barcode", meta, report)):
                        with mock.patch.dict(os.environ, {"SCAN2_MODEL_BARCODE": "0"}, clear=False):
                            scan2.main(
                                model_dir=model_dir,
                                sn_dir=sn_dir,
                                out_jsonl=os.path.join(root, "out.jsonl"),
                                debug_log=os.path.join(root, "debug.log"),
                            )
            finally:
                scan2.set_log_sink(old_sink)

            joined = "\n".join(logs)
            self.assertIn("[条码完成] a__label_1 SN -> 命中 4E25B0105849（扫描条形码）", joined)
            self.assertNotIn("[条码开始] a__label_1 SN", joined)

    def test_set_log_sink_is_forwarded_to_loaded_ocr_module(self):
        scan2 = _import_scan2()

        calls = []
        fake_ocr = types.SimpleNamespace(set_log_sink=calls.append)
        old_module = scan2.OCR_MODULE
        old_sink = scan2.LOG_SINK
        scan2.OCR_MODULE = fake_ocr
        try:
            scan2.set_log_sink("sink-a")
            self.assertEqual(calls, ["sink-a"])
        finally:
            scan2.OCR_MODULE = old_module
            scan2.set_log_sink(old_sink)

    def test_label_crop_barcode_is_used_when_sn_crop_is_missing(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            label_path = os.path.join(root, "label.png")
            open(label_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": "a__label_1", "label_crop": label_path}) + "\n")

            out_jsonl = os.path.join(root, "out.jsonl")
            with mock.patch.object(scan2, "read_barcodes", return_value=["SN:4E25A0170000"]):
                with mock.patch.object(scan2, "load_for_ocr_color") as load_ocr:
                    with mock.patch.dict(os.environ, {"SCAN2_SCAN_LABEL_WITHOUT_SN": "1"}, clear=False):
                        stats = scan2.main(
                            model_dir=model_dir,
                            sn_dir=sn_dir,
                            out_jsonl=out_jsonl,
                            debug_log=os.path.join(root, "debug.log"),
                        )

            with open(out_jsonl, "r", encoding="utf-8") as f:
                row = json.loads(f.readline())
            self.assertEqual(row["sn"], "4E25A0170000")
            self.assertEqual(row["sn_src"], "barcode")
            self.assertEqual(stats["sn_total"], 1)
            self.assertEqual(stats["sn_success"], 1)
            load_ocr.assert_not_called()

    def test_model_barcode_is_enabled_by_default(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            model_path = os.path.join(model_dir, "a__label_1__model.png")
            open(model_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": "a__label_1", "model_path": model_path}) + "\n")

            with mock.patch.object(scan2, "recognize_model_barcode", return_value=("AP162E", "raw", "barcode")) as recognize_model:
                with mock.patch.object(scan2, "recognize_model_ocr") as recognize_model_ocr:
                    with mock.patch.dict(os.environ, {}, clear=True):
                        stats = scan2.main(
                            model_dir=model_dir,
                            sn_dir=sn_dir,
                            out_jsonl=os.path.join(root, "out.jsonl"),
                            debug_log=os.path.join(root, "debug.log"),
                        )

            recognize_model.assert_called_once()
            self.assertEqual(recognize_model.call_args.kwargs["label_id"], "a__label_1")
            recognize_model_ocr.assert_not_called()
            self.assertEqual(stats["model_total"], 1)
            self.assertEqual(stats["model_success"], 1)
            self.assertEqual(stats["model_barcode_hits"], 1)
            self.assertEqual(stats["model_barcode_hit_rate"], 1.0)

    def test_model_barcode_can_be_disabled_with_env_flag(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            model_path = os.path.join(model_dir, "a__label_1__model.png")
            open(model_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": "a__label_1", "model_path": model_path}) + "\n")

            with mock.patch.object(scan2, "recognize_model_barcode") as recognize_model:
                with mock.patch.object(scan2, "recognize_model_ocr", return_value=("MODEL1", "raw", "test")):
                    with mock.patch.dict(os.environ, {"SCAN2_MODEL_BARCODE": "0"}, clear=True):
                        scan2.main(
                            model_dir=model_dir,
                            sn_dir=sn_dir,
                            out_jsonl=os.path.join(root, "out.jsonl"),
                            debug_log=os.path.join(root, "debug.log"),
                        )

            recognize_model.assert_not_called()

    def test_scan_worker_count_auto_and_env_overrides(self):
        scan2 = _import_scan2()

        with mock.patch.object(scan2.os, "cpu_count", return_value=16):
            with mock.patch.dict(os.environ, {}, clear=True):
                self.assertEqual(scan2.scan_worker_count("barcode"), 8)
                self.assertEqual(scan2.scan_worker_count("ocr"), 1)

            with mock.patch.dict(os.environ, {"SCAN2_WORKERS": "3"}, clear=True):
                self.assertEqual(scan2.scan_worker_count("barcode"), 3)
                self.assertEqual(scan2.scan_worker_count("ocr"), 3)

            with mock.patch.dict(os.environ, {"SCAN2_OCR_WORKERS": "4", "SCAN2_WORKERS": "3"}, clear=True):
                self.assertEqual(scan2.scan_worker_count("barcode"), 3)
                self.assertEqual(scan2.scan_worker_count("ocr"), 4)

            with mock.patch.dict(os.environ, {"SCAN2_PARALLEL": "0", "SCAN2_WORKERS": "5"}, clear=True):
                self.assertEqual(scan2.scan_worker_count("barcode"), 1)
                self.assertEqual(scan2.scan_worker_count("ocr"), 1)

    def test_main_runs_barcode_batch_before_ocr_fallback_and_preserves_rows(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            rows = []
            for label_id in ("b__label_1", "a__label_1"):
                model_path = os.path.join(model_dir, f"{label_id}__model.png")
                open(model_path, "wb").close()
                rows.append({"label_id": label_id, "model_path": model_path})
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                for row in rows:
                    manifest.write(json.dumps(row) + "\n")

            events = []

            def fake_barcode(_path, label_id=""):
                events.append(f"barcode:{label_id}")
                return "", "", "barcode_no_match"

            def fake_ocr(_path, label_id="", verify_barcode_visual=False):
                events.append(f"ocr:{label_id}")
                return "AP162E", "raw", "ocr_file"

            out_jsonl = os.path.join(root, "out.jsonl")
            with mock.patch.object(scan2, "recognize_model_barcode", side_effect=fake_barcode):
                with mock.patch.object(scan2, "recognize_model_ocr", side_effect=fake_ocr):
                    with mock.patch.dict(
                        os.environ,
                        {
                            "SCAN2_BARCODE_WORKERS": "2",
                            "SCAN2_OCR_WORKERS": "2",
                            "SCAN2_OCR_FALLBACK": "1",
                        },
                        clear=False,
                    ):
                        stats = scan2.main(
                            model_dir=model_dir,
                            sn_dir=sn_dir,
                            out_jsonl=out_jsonl,
                            debug_log=os.path.join(root, "debug.log"),
                        )

            last_barcode = max(i for i, value in enumerate(events) if value.startswith("barcode:"))
            first_ocr = min(i for i, value in enumerate(events) if value.startswith("ocr:"))
            self.assertLess(last_barcode, first_ocr)
            with open(out_jsonl, "r", encoding="utf-8") as f:
                output_rows = [json.loads(line) for line in f if line.strip()]
            self.assertEqual([row["label_id"] for row in output_rows], ["a__label_1", "b__label_1"])
            self.assertEqual(stats["model_total"], 2)
            self.assertEqual(stats["model_ocr_recoveries"], 2)

    def test_model_recognition_skips_barcode_cli_when_disabled(self):
        scan2 = _import_scan2()
        fake_img = object()

        with mock.patch.object(scan2, "try_model_from_barcode") as barcode_mock:
            with mock.patch.object(scan2, "load_for_ocr_color", return_value=None):
                with mock.patch.object(scan2, "load_and_preprocess", return_value=fake_img):
                    with mock.patch.object(scan2, "ocr_text_with_details", return_value=("", "", [])):
                        model, _raw, source = scan2.recognize_model("model.png", use_barcode=False)

        self.assertEqual(model, "")
        self.assertEqual(source, "none")
        barcode_mock.assert_not_called()

    def test_main_uses_sn_hint_before_model_ocr(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            label_id = "a__label_1"
            model_path = os.path.join(model_dir, f"{label_id}__model.png")
            sn_path = os.path.join(sn_dir, f"{label_id}__sn.png")
            open(model_path, "wb").close()
            open(sn_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(json.dumps({"label_id": label_id, "model_path": model_path, "sn_path": sn_path}) + "\n")

            sn_meta = {
                "barcode_found": True,
                "ocr_text_found": False,
                "barcode_status": "hit",
                "barcode_attempts": 1,
                "barcode_decoded_count": 1,
            }
            with mock.patch.object(scan2, "recognize_model_barcode", return_value=("", "", "barcode_no_match")):
                with mock.patch.object(
                    scan2,
                    "_recognize_sn_barcode",
                    return_value=("4E2640069362", "raw", "barcode", sn_meta, None),
                ):
                    with mock.patch.object(scan2, "recognize_model_ocr") as model_ocr:
                        stats = scan2.main(
                            model_dir=model_dir,
                            sn_dir=sn_dir,
                            out_jsonl=os.path.join(root, "out.jsonl"),
                            debug_log=os.path.join(root, "debug.log"),
                        )

            model_ocr.assert_not_called()
            with open(os.path.join(root, "out.jsonl"), "r", encoding="utf-8") as f:
                row = json.loads(f.readline())
            self.assertEqual(row["model"], "S380-S8P2T")
            self.assertEqual(row["model_src"], "sn_hint")
            self.assertEqual(stats["model_ocr_recoveries"], 0)

    def test_main_falls_back_to_label_ocr_when_model_crop_ocr_misses(self):
        scan2 = _import_scan2()

        with tempfile.TemporaryDirectory() as root:
            stage2 = os.path.join(root, "stage2_fields")
            model_dir = os.path.join(stage2, "model")
            sn_dir = os.path.join(stage2, "sn")
            os.makedirs(model_dir)
            os.makedirs(sn_dir)
            label_id = "a__label_1"
            model_path = os.path.join(model_dir, f"{label_id}__model.png")
            label_path = os.path.join(root, f"{label_id}.png")
            open(model_path, "wb").close()
            open(label_path, "wb").close()
            with open(os.path.join(stage2, "manifest.jsonl"), "w", encoding="utf-8") as manifest:
                manifest.write(
                    json.dumps(
                        {
                            "label_id": label_id,
                            "model_path": model_path,
                            "label_crop": label_path,
                        }
                    )
                    + "\n"
                )

            sn_meta = {
                "barcode_found": False,
                "ocr_text_found": False,
                "barcode_status": "miss",
                "barcode_attempts": 1,
                "barcode_decoded_count": 0,
            }
            sn_report = types.SimpleNamespace(status="miss", results=[])
            with mock.patch.object(scan2, "recognize_model_barcode", return_value=("", "", "barcode_no_match")):
                with mock.patch.object(
                    scan2,
                    "_recognize_sn_barcode",
                    return_value=("", "", "missing", sn_meta, sn_report),
                ):
                    with mock.patch.object(scan2, "recognize_model_ocr", return_value=("", "noise", "none")) as model_ocr:
                        with mock.patch.object(
                            scan2,
                            "recognize_model_label_ocr",
                            return_value=("AP162E", "Model: AP162E", "ocr_label"),
                        ) as label_ocr:
                            stats = scan2.main(
                                model_dir=model_dir,
                                sn_dir=sn_dir,
                                out_jsonl=os.path.join(root, "out.jsonl"),
                                debug_log=os.path.join(root, "debug.log"),
                            )

            model_ocr.assert_called_once()
            label_ocr.assert_called_once_with(label_path, label_id=label_id)
            with open(os.path.join(root, "out.jsonl"), "r", encoding="utf-8") as f:
                row = json.loads(f.readline())
            self.assertEqual(row["model"], "AP162E")
            self.assertEqual(row["model_src"], "ocr_label")
            self.assertEqual(stats["model_ocr_recoveries"], 1)

    def test_model_recognition_prefers_file_path_ocr(self):
        scan2 = _import_scan2()

        with mock.patch.object(
            scan2,
            "ocr_text_with_details",
            return_value=(
                "Model: S380-S8P2T",
                "Model:S380-S8P2T",
                [{"text": "Model: S380-S8P2T", "score": 0.99}],
            ),
        ) as ocr_mock:
            with mock.patch.object(scan2, "load_for_ocr_color") as color_mock:
                with mock.patch.object(scan2, "load_and_preprocess") as bin_mock:
                    model, raw, source = scan2.recognize_model("model.png", use_barcode=False)

        self.assertEqual(model, "S380-S8P2T")
        self.assertEqual(raw, "Model: S380-S8P2T")
        self.assertEqual(source, "ocr_file")
        ocr_mock.assert_called_once_with("model.png")
        color_mock.assert_not_called()
        bin_mock.assert_not_called()

    def test_model_ocr_result_prefers_spaced_text_over_concat(self):
        scan2 = _import_scan2()

        model = scan2.extract_model_from_ocr_result(
            "Model: AP162E 9SC:AD1625(1",
            "Model: AP162E9SC:AD1625(1",
        )

        self.assertEqual(model, "AP162E")

    def test_s380_s8p2t_ocr_noise_is_normalized(self):
        scan2 = _import_scan2()

        self.assertEqual(scan2.extract_model_from_text("MO8S-0802 Wac"), "S380-S8P2T")
        self.assertEqual(scan2.extract_model_from_text("M08S-0802 Wac"), "S380-S8P2T")


class CropTempFileTest(unittest.TestCase):
    def test_original_path_for_label_id_resolves_current_input_dir(self):
        crop = _import_crop()
        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "input")
            os.makedirs(input_dir)
            original = os.path.join(input_dir, "image_01.jpg")
            with open(original, "wb") as f:
                f.write(b"image")

            crop.configure_paths(input_dir=input_dir, out_dir=root)

            self.assertEqual(
                crop.original_path_for_label_id("image_01.jpg__label_3"),
                original,
            )

    def test_recursive_input_paths_get_safe_label_id_and_original_mapping(self):
        crop = _import_crop()
        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "input")
            nested_dir = os.path.join(input_dir, "2026.2.2")
            os.makedirs(nested_dir)
            nested = os.path.join(nested_dir, "same.jpg")
            top_level = os.path.join(input_dir, "top.jpg")
            for path in (nested, top_level):
                with open(path, "wb") as f:
                    f.write(b"image")

            crop.configure_paths(input_dir=input_dir, out_dir=root)
            self.assertEqual(crop.list_images(input_dir), [nested, top_level])

            nested_base = crop.input_label_base(nested)
            self.assertNotIn(os.sep, nested_base)
            self.assertIn("2026.2.2", nested_base)
            self.assertEqual(crop.original_path_for_label_id(f"{nested_base}__label_1"), nested)
            self.assertEqual(crop.input_label_base(top_level), "top.jpg")

    def test_stage1_uses_extension_in_label_name_to_avoid_same_stem_collision(self):
        crop = _import_crop()
        fake_img = np.zeros((100, 100, 3), dtype=np.uint8)
        pred = {"x": 50, "y": 50, "width": 20, "height": 20, "class": crop.MODEL1_LABEL_CLASS}

        with tempfile.TemporaryDirectory() as root:
            crop.configure_paths(input_dir=os.path.join(root, "input"), out_dir=root)
            os.makedirs(crop.INPUT_DIR)
            path_png = os.path.join(root, "a", "same.png")
            path_jpg = os.path.join(root, "b", "same.jpg")

            with mock.patch.object(crop, "read_image", return_value=fake_img):
                with mock.patch.object(crop, "infer_with_resize", return_value=[pred]):
                    with mock.patch.object(crop, "crop_from_pred", return_value=fake_img):
                        with mock.patch.object(crop, "stage1_is_product_label_candidate_crop", return_value=True):
                            with mock.patch.object(crop, "stage1_tighten_label_crop", return_value=fake_img):
                                with mock.patch.object(crop, "stage1_is_product_label_crop", return_value=True):
                                    with mock.patch.object(crop, "save_png_required", side_effect=lambda path, _img, _ctx: path):
                                        with mock.patch.object(crop, "stage1_save_preview_enabled", return_value=False):
                                            out_png = crop.stage1_crop_labels(path_png)
                                            out_jpg = crop.stage1_crop_labels(path_jpg)

            self.assertEqual(os.path.basename(out_png[0]), "same.png__label_1.png")
            self.assertEqual(os.path.basename(out_jpg[0]), "same.jpg__label_1.png")
            self.assertNotEqual(os.path.basename(out_png[0]), os.path.basename(out_jpg[0]))

    def test_main_removes_stage1_orphan_when_stage2_rejects_label(self):
        crop = _import_crop()
        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "input")
            out_dir = os.path.join(root, "out")
            os.makedirs(input_dir)
            img_path = os.path.join(input_dir, "image_01.jpg")
            with open(img_path, "wb") as f:
                f.write(b"image")

            def fake_stage1(_img_path):
                label_path = os.path.join(crop.STAGE1_DIR, "image_01.jpg__label_1.png")
                os.makedirs(os.path.dirname(label_path), exist_ok=True)
                with open(label_path, "wb") as f:
                    f.write(b"label")
                return [label_path]

            with mock.patch.object(crop, "crop_worker_count", return_value=1):
                with mock.patch.object(crop, "stage1_crop_labels", side_effect=fake_stage1):
                    with mock.patch.object(crop, "stage2_crop_fields", return_value=None):
                        with mock.patch.object(crop, "stage1_keep_all_crops_enabled", return_value=False):
                            stats = crop.main(input_dir=input_dir, out_dir=out_dir, clean=True, log_level="error")

            label_path = os.path.join(out_dir, "stage1_labels", "image_01.jpg__label_1.png")
            self.assertFalse(os.path.exists(label_path))
            self.assertEqual(stats["label_count"], 0)
            self.assertEqual(stats["manifest_rows"], 0)

    def test_main_emits_realtime_stage2_progress_logs(self):
        crop = _import_crop()
        logs = []
        with tempfile.TemporaryDirectory() as root:
            input_dir = os.path.join(root, "input")
            out_dir = os.path.join(root, "out")
            os.makedirs(input_dir)
            img_path = os.path.join(input_dir, "image_01.jpg")
            with open(img_path, "wb") as f:
                f.write(b"image")

            label_paths = [
                os.path.join(out_dir, "stage1_labels", "image_01.jpg__label_1.png"),
                os.path.join(out_dir, "stage1_labels", "image_01.jpg__label_2.png"),
            ]

            def fake_stage1(_img_path):
                for label_path in label_paths:
                    os.makedirs(os.path.dirname(label_path), exist_ok=True)
                    with open(label_path, "wb") as f:
                        f.write(b"label")
                return list(label_paths)

            def fake_stage2(label_path):
                if label_path.endswith("__label_1.png"):
                    return {
                        "label_id": "image_01.jpg__label_1",
                        "label_crop": label_path,
                        "model_path": "model.png",
                        "sn_path": "sn.png",
                        "part_no_path": None,
                    }
                return None

            old_sink = crop.LOG_SINK
            crop.set_log_sink(logs.append)
            with mock.patch.dict(os.environ, {"CROP_PROGRESS_LOG": "1"}):
                with mock.patch.object(crop, "crop_worker_count", return_value=1):
                    with mock.patch.object(crop, "stage1_crop_labels", side_effect=fake_stage1):
                        with mock.patch.object(crop, "stage2_crop_fields", side_effect=fake_stage2):
                            stats = crop.main(input_dir=input_dir, out_dir=out_dir, clean=True, log_level="info")
            crop.set_log_sink(old_sink)

            joined = "\n".join(logs)
            self.assertEqual(stats["label_count"], 1)
            self.assertIn("[2/4][字段完成] image_01.jpg__label_1 -> 命中 Model/SN", joined)
            self.assertIn("[2/4][字段完成] image_01.jpg__label_2 -> 未保留", joined)
            self.assertNotIn("[2/4][字段开始] image_01.jpg__label_1", joined)
            self.assertNotIn("[2/4][字段开始] image_01.jpg__label_2", joined)

    def test_save_png_required_raises_when_write_fails(self):
        crop = _import_crop()
        with tempfile.TemporaryDirectory() as root:
            target = os.path.join(root, "out.png")
            with mock.patch.object(crop, "save_png", return_value=False):
                with self.assertRaisesRegex(RuntimeError, "Failed to write test crop"):
                    crop.save_png_required(target, object(), "test crop")

    def test_infer_with_resize_uses_unique_temp_file_and_cleans_it(self):
        crop = _import_crop()

        calls = []

        class FakeClient:
            def infer(self, path, model_id):
                calls.append(path)
                self.seen_exists = os.path.exists(path)
                return {"predictions": []}

        fake_client = FakeClient()
        fake_img = types.SimpleNamespace(shape=(100, 200, 3))

        def write_tmp(_bgr, path, quality=85):
            with open(path, "wb") as f:
                f.write(b"tmp")
            return True

        with tempfile.TemporaryDirectory() as root:
            crop.TMP_DIR = root
            with mock.patch.object(crop, "get_inference_client", return_value=fake_client):
                with mock.patch.object(crop, "_write_tmp_jpg", side_effect=write_tmp):
                    self.assertEqual(crop.infer_with_resize(fake_img, "same__name.png", "model/1"), [])
                    self.assertEqual(crop.infer_with_resize(fake_img, "same__name.png", "model/1"), [])

            self.assertEqual(len(calls), 2)
            self.assertNotEqual(calls[0], calls[1])
            self.assertFalse(os.path.exists(calls[0]))
            self.assertFalse(os.path.exists(calls[1]))

    def test_crop_worker_count_defaults_parallel_and_respects_overrides(self):
        crop = _import_crop()

        with mock.patch.dict(os.environ, {}, clear=False):
            for key in ("CROP_WORKERS", "CROP_STAGE1_WORKERS", "CROP_STAGE2_WORKERS"):
                os.environ.pop(key, None)
            os.environ.pop("CROP_INFERENCE_BACKEND", None)

            with mock.patch.object(crop.os, "cpu_count", return_value=16):
                with mock.patch.object(crop, "_local_yolo_cuda_available", return_value=False):
                    self.assertEqual(crop.inference_backend(), "local")
                    self.assertEqual(crop.crop_worker_count("stage2"), 2)

                with mock.patch.object(crop, "_local_yolo_cuda_available", return_value=True):
                    self.assertEqual(crop.crop_worker_count("stage2"), 4)

                os.environ["CROP_INFERENCE_BACKEND"] = "roboflow"
                self.assertEqual(crop.crop_worker_count("stage2"), 8)

                os.environ["CROP_INFERENCE_BACKEND"] = "local"
                with mock.patch.object(crop, "_local_yolo_cuda_available", return_value=False):
                    self.assertEqual(crop.crop_worker_count("stage2"), 2)

                os.environ["CROP_INFERENCE_BACKEND"] = "auto"
                with mock.patch.object(crop, "_local_yolo_cuda_available", return_value=False):
                    self.assertEqual(crop.crop_worker_count("stage2"), 2)

            os.environ["CROP_WORKERS"] = "2"
            self.assertEqual(crop.crop_worker_count("stage2"), 2)

            os.environ["CROP_STAGE2_WORKERS"] = "3"
            self.assertEqual(crop.crop_worker_count("stage2"), 3)

    def test_map_ordered_preserves_input_order_when_parallel(self):
        crop = _import_crop()

        def worker(value):
            if value == 1:
                time.sleep(0.02)
            return value * 10

        self.assertEqual(crop._map_ordered([3, 1, 2], worker, workers=3), [30, 10, 20])

    def test_infer_with_resize_reports_cloudflare_block_without_html(self):
        crop = _import_crop()

        class Cloudflare403(Exception):
            status_code = 403
            api_message = "<!DOCTYPE html><title>Attention Required! | Cloudflare</title><h1>Sorry, you have been blocked</h1>"
            description = "403 Client Error: Forbidden"

        class FakeClient:
            def __init__(self):
                self.calls = 0

            def infer(self, path, model_id):
                self.calls += 1
                raise Cloudflare403()

        fake_client = FakeClient()
        fake_img = types.SimpleNamespace(shape=(100, 200, 3))

        def write_tmp(_bgr, path, quality=85):
            with open(path, "wb") as f:
                f.write(b"tmp")
            return True

        with tempfile.TemporaryDirectory() as root:
            crop.TMP_DIR = root
            with mock.patch.dict(os.environ, {"CROP_INFERENCE_BACKEND": "roboflow"}):
                with mock.patch.object(crop, "get_inference_client", return_value=fake_client):
                    with mock.patch.object(crop, "_write_tmp_jpg", side_effect=write_tmp):
                        with self.assertRaises(RuntimeError) as ctx:
                            crop.infer_with_resize(fake_img, "image.png", "huawei-2ha7t/7")

            message = str(ctx.exception)
            self.assertIn("Cloudflare", message)
            self.assertIn("HTTP 403", message)
            self.assertNotIn("<!DOCTYPE", message)
            self.assertNotIn("Sorry, you have been blocked", message)
            self.assertEqual(fake_client.calls, 1)

    def test_auto_backend_falls_back_to_local_on_cloudflare_block(self):
        crop = _import_crop()
        crop.auto_fallback_backend = None

        class Cloudflare403(Exception):
            status_code = 403
            api_message = "<html>Cloudflare: you have been blocked</html>"
            description = "403 Client Error: Forbidden"

        class CloudClient:
            def __init__(self):
                self.calls = 0

            def infer(self, path, model_id):
                self.calls += 1
                raise Cloudflare403()

        class LocalClient:
            def __init__(self):
                self.calls = 0

            def infer(self, path, model_id):
                self.calls += 1
                return {"predictions": [{"x": 10, "y": 20, "width": 30, "height": 40}]}

        cloud_client = CloudClient()
        local_client = LocalClient()
        requested_backends = []
        fake_img = types.SimpleNamespace(shape=(100, 200, 3))

        def get_client(backend=None):
            requested_backends.append(backend or crop.inference_backend())
            return local_client if backend == "local" else cloud_client

        def write_tmp(_bgr, path, quality=85):
            with open(path, "wb") as f:
                f.write(b"tmp")
            return True

        with tempfile.TemporaryDirectory() as root:
            crop.TMP_DIR = root
            with mock.patch.dict(os.environ, {"CROP_INFERENCE_BACKEND": "auto"}):
                with mock.patch.object(crop, "get_inference_client", side_effect=get_client):
                    with mock.patch.object(crop, "_write_tmp_jpg", side_effect=write_tmp):
                        preds = crop.infer_with_resize(fake_img, "image.png", "huawei-2ha7t/7")

        self.assertEqual(preds, [{"x": 10, "y": 20, "width": 30, "height": 40}])
        self.assertEqual(requested_backends, ["roboflow", "local"])
        self.assertEqual(cloud_client.calls, 1)
        self.assertEqual(local_client.calls, 1)
        self.assertEqual(crop.auto_fallback_backend, "local")

    def test_auto_backend_reuses_local_after_cloudflare_block(self):
        crop = _import_crop()
        crop.auto_fallback_backend = None

        class Cloudflare403(Exception):
            status_code = 403
            api_message = "<html>Attention Required! | Cloudflare</html>"
            description = "403 Client Error: Forbidden"

        class CloudClient:
            def __init__(self):
                self.calls = 0

            def infer(self, path, model_id):
                self.calls += 1
                raise Cloudflare403()

        class LocalClient:
            def __init__(self):
                self.calls = 0

            def infer(self, path, model_id):
                self.calls += 1
                return {"predictions": []}

        cloud_client = CloudClient()
        local_client = LocalClient()
        requested_backends = []
        fake_img = types.SimpleNamespace(shape=(100, 200, 3))

        def get_client(backend=None):
            requested_backends.append(backend or crop.inference_backend())
            return local_client if backend == "local" else cloud_client

        def write_tmp(_bgr, path, quality=85):
            with open(path, "wb") as f:
                f.write(b"tmp")
            return True

        with tempfile.TemporaryDirectory() as root:
            crop.TMP_DIR = root
            with mock.patch.dict(os.environ, {"CROP_INFERENCE_BACKEND": "auto"}):
                with mock.patch.object(crop, "get_inference_client", side_effect=get_client):
                    with mock.patch.object(crop, "_write_tmp_jpg", side_effect=write_tmp):
                        self.assertEqual(
                            crop.infer_with_resize(fake_img, "image-a.png", "huawei-2ha7t/7"),
                            [],
                        )
                        self.assertEqual(
                            crop.infer_with_resize(fake_img, "image-b.png", "sn_model/9"),
                            [],
                        )

        self.assertEqual(requested_backends, ["roboflow", "local", "local"])
        self.assertEqual(cloud_client.calls, 1)
        self.assertEqual(local_client.calls, 2)

    def test_dotenv_loads_from_packaged_internal_dir(self):
        crop = _import_crop()

        with tempfile.TemporaryDirectory() as root:
            internal = os.path.join(root, "_internal")
            os.makedirs(internal)
            with open(os.path.join(internal, ".env"), "w", encoding="utf-8-sig") as f:
                f.write("API_KEY=packaged-test-key\n")

            old_frozen = getattr(sys, "frozen", None)
            old_executable = sys.executable
            old_api_key = os.environ.pop("API_KEY", None)
            try:
                sys.frozen = True
                sys.executable = os.path.join(root, "HuaweiOCR.exe")
                with mock.patch("os.getcwd", return_value=root):
                    crop.load_dotenv()
                self.assertEqual(os.environ.get("API_KEY"), "packaged-test-key")
            finally:
                if old_frozen is None:
                    try:
                        delattr(sys, "frozen")
                    except AttributeError:
                        pass
                else:
                    sys.frozen = old_frozen
                sys.executable = old_executable
                if old_api_key is None:
                    os.environ.pop("API_KEY", None)
                else:
                    os.environ["API_KEY"] = old_api_key

    def test_dotenv_empty_value_does_not_block_packaged_internal_dir(self):
        crop = _import_crop()

        with tempfile.TemporaryDirectory() as root:
            internal = os.path.join(root, "_internal")
            os.makedirs(internal)
            with open(os.path.join(root, ".env"), "w", encoding="utf-8") as f:
                f.write("API_KEY=\n")
            with open(os.path.join(internal, ".env"), "w", encoding="utf-8") as f:
                f.write("API_KEY=packaged-test-key\n")

            old_frozen = getattr(sys, "frozen", None)
            old_executable = sys.executable
            old_api_key = os.environ.pop("API_KEY", None)
            try:
                sys.frozen = True
                sys.executable = os.path.join(root, "HuaweiOCR.exe")
                with mock.patch("os.getcwd", return_value=root):
                    crop.load_dotenv()
                self.assertEqual(os.environ.get("API_KEY"), "packaged-test-key")
            finally:
                if old_frozen is None:
                    try:
                        delattr(sys, "frozen")
                    except AttributeError:
                        pass
                else:
                    sys.frozen = old_frozen
                sys.executable = old_executable
                if old_api_key is None:
                    os.environ.pop("API_KEY", None)
                else:
                    os.environ["API_KEY"] = old_api_key


class GuiPipelineTest(unittest.TestCase):
    def test_same_basename_sources_are_staged_with_unique_names(self):
        import gui_pipeline

        with tempfile.TemporaryDirectory() as root:
            source_a = os.path.join(root, "a")
            source_b = os.path.join(root, "b")
            os.makedirs(source_a)
            os.makedirs(source_b)
            path_a = os.path.join(source_a, "same.png")
            path_b = os.path.join(source_b, "same.png")
            with open(path_a, "wb") as f:
                f.write(b"a")
            with open(path_b, "wb") as f:
                f.write(b"b")

            run_dir, records = gui_pipeline.copy_images_to_unique_run_dir(
                [path_a, path_b],
                os.path.join(root, "new_images"),
            )

            names = [record["input_name"] for record in records]
            self.assertEqual(len(names), 2)
            self.assertEqual(len(set(names)), 2)
            self.assertEqual(names, ["same.png", "same_2.png"])
            self.assertTrue(os.path.exists(os.path.join(run_dir, names[0])))
            self.assertTrue(os.path.exists(os.path.join(run_dir, names[1])))
            with open(os.path.join(run_dir, "source_manifest.jsonl"), "r", encoding="utf-8") as f:
                manifest_rows = [json.loads(line) for line in f]
            manifest_text = json.dumps(manifest_rows, ensure_ascii=False)
            self.assertNotIn("source_path", manifest_text)
            self.assertNotIn("input_path", manifest_text)
            self.assertNotIn(os.path.abspath(root), manifest_text)
            for row in manifest_rows:
                self.assertEqual(set(row), {"source_index", "input_name", "sha256"})
            self.assertEqual([row["source_index"] for row in manifest_rows], [1, 2])
            self.assertEqual([row["input_name"] for row in manifest_rows], names)
            self.assertTrue(all(len(row["sha256"]) == 64 for row in manifest_rows))

    def test_display_sources_are_chinese(self):
        sys.modules.pop("gui_app", None)
        sys.modules.pop("numpy", None)
        sys.modules.pop("app_paths", None)
        sys.modules.pop("huaweiocr.io.paths_runtime", None)
        import gui_app

        self.assertEqual(gui_app._display_model_src("barcode"), "扫描条形码")
        self.assertEqual(gui_app._display_model_src("ocr_file"), "文字识别")
        self.assertEqual(gui_app._display_sn_src("barcode_decoder_miss"), "未扫到条形码")

    def test_gui_import_does_not_import_pipeline_modules(self):
        sys.modules.pop("gui_app", None)
        sys.modules.pop("crop", None)
        sys.modules.pop("scan2", None)
        sys.modules.pop("barcode", None)
        sys.modules.pop("huaweiocr.barcode.generic", None)
        sys.modules.pop("cv2", None)
        sys.modules.pop("numpy", None)
        sys.modules.pop("pyzbar", None)
        sys.modules.pop("pyzbar.pyzbar", None)

        importlib.import_module("gui_app")

        self.assertNotIn("crop", sys.modules)
        self.assertNotIn("scan2", sys.modules)

    def test_gui_log_mask_keeps_app_relative_output_paths(self):
        sys.modules.pop("gui_app", None)
        import gui_app

        app_path = os.path.join(os.getcwd(), "stage2_fields", "manifest.jsonl")
        external_path = r"F:\wechat\xwechat_files\sample.jpg"

        masked = gui_app._mask_path_text(
            f"Manifest: {app_path}; source: {external_path}"
        )

        self.assertIn(os.path.join("stage2_fields", "manifest.jsonl"), masked)
        self.assertIn("source: [path]", masked)
        self.assertNotIn(external_path, masked)

    def test_gui_write_log_appends_directly_on_main_thread(self):
        sys.modules.pop("gui_app", None)
        import gui_app

        captured = []
        after_calls = []
        app = types.SimpleNamespace(
            _main_thread_ident=threading.get_ident(),
            _append_log_lines=lambda lines: captured.extend(lines),
            after=lambda delay, callback=None: after_calls.append((delay, callback)),
        )

        gui_app.App.write_log(app, "主线程日志")

        self.assertEqual(captured, ["主线程日志"])
        self.assertEqual(after_calls, [])

    def test_gui_write_log_batches_background_updates(self):
        sys.modules.pop("gui_app", None)
        import gui_app

        captured = []
        scheduled = []
        app = types.SimpleNamespace()
        app._main_thread_ident = threading.get_ident()
        app._log_queue = queue.SimpleQueue()
        app._log_poll_interval_ms = 50
        app._append_log_lines = lambda lines: captured.extend(lines)
        app.after = lambda delay, callback=None: scheduled.append((delay, callback))
        app._poll_log_queue = lambda: gui_app.App._poll_log_queue(app)
        app._flush_log_buffer = lambda: gui_app.App._flush_log_buffer(app)

        fake_bg_ident = app._main_thread_ident + 1
        with mock.patch("threading.get_ident", return_value=fake_bg_ident):
            gui_app.App.write_log(app, "后台日志1")
            gui_app.App.write_log(app, "后台日志2")

        self.assertEqual(len(scheduled), 0)

        gui_app.App._flush_log_buffer(app)

        self.assertEqual(captured, ["后台日志1", "后台日志2"])
        self.assertEqual(len(scheduled), 1)

    def _run_gui_pipeline_with_fakes(self, module_name):
        sys.modules.pop(module_name, None)
        gui_module = importlib.import_module(module_name)

        with tempfile.TemporaryDirectory() as root:
            input_root = os.path.join(root, "new_images")
            run_dir = os.path.join(input_root, "gui_run_test")
            stage2_dir = os.path.join(run_dir, "stage2_fields")
            os.makedirs(run_dir)

            crop_calls = []
            scan2_calls = []

            crop_module = types.SimpleNamespace(
                DEFAULT_INPUT_DIR=input_root,
                STAGE1_DIR=os.path.join(run_dir, "stage1_labels"),
                STAGE2_DIR=stage2_dir,
                OUT_MODEL_DIR=os.path.join(stage2_dir, "model"),
                OUT_SN_DIR=os.path.join(stage2_dir, "sn"),
                LOG_SINK=None,
            )

            def crop_main(**kwargs):
                crop_calls.append(kwargs)
                out_dir = kwargs.get("out_dir") or run_dir
                crop_module.STAGE1_DIR = os.path.join(out_dir, "stage1_labels")
                crop_module.STAGE2_DIR = os.path.join(out_dir, "stage2_fields")
                crop_module.OUT_MODEL_DIR = os.path.join(crop_module.STAGE2_DIR, "model")
                crop_module.OUT_SN_DIR = os.path.join(crop_module.STAGE2_DIR, "sn")
                return {"label_count": 1, "manifest_rows": 1}

            crop_module.main = crop_main
            crop_module.set_log_sink = lambda sink: setattr(crop_module, "LOG_SINK", sink)

            scan2_module = types.SimpleNamespace(
                OUT_JSONL=os.path.join(stage2_dir, "model_sn_ocr.jsonl"),
                LOG_SINK=None,
            )

            def scan2_main(**kwargs):
                scan2_calls.append(kwargs)
                scan2_module.OUT_JSONL = kwargs["out_jsonl"]
                return {"rows": 1}

            scan2_module.main = scan2_main
            scan2_module.set_log_sink = lambda sink: setattr(scan2_module, "LOG_SINK", sink)

            app = types.SimpleNamespace(
                image_paths=[os.path.join(root, "source.png")],
                btn_start=types.SimpleNamespace(config=lambda **_kwargs: None),
                write_log=lambda _text: None,
                after=lambda _delay, callback=None: callback() if callback else None,
                load_results_into_table=lambda: None,
                _format_issue_summary=lambda: "ok",
                strings=gui_module.get_strings("zh"),
            )

            with mock.patch.object(gui_module, "load_pipeline_modules", return_value=(crop_module, scan2_module)):
                with mock.patch.object(
                    gui_module,
                    "copy_images_to_unique_run_dir",
                    return_value=(run_dir, [{"input_name": "source.png"}]),
                ) as copy_run_dir:
                    gui_module.App.run_pipeline(app)

            return {
                "input_root": input_root,
                "run_dir": run_dir,
                "image_paths": app.image_paths,
                "copy_call": copy_run_dir.call_args,
                "crop_call": crop_calls[0],
                "scan2_call": scan2_calls[0],
            }

    def test_gui_en_entry_delegates_to_gui_app(self):
        sys.modules.pop("gui_app_en", None)
        gui_app_en = importlib.import_module("gui_app_en")
        self.assertEqual(gui_app_en.main.__module__, "gui_app")

    def test_gui_run_pipeline_uses_unique_run_dir_as_crop_out_dir(self):
        # gui_app_en 自 i18n 合并后是 gui_app 的薄壳，测 gui_app 即覆盖两个入口。
        for module_name in ("gui_app",):
            with self.subTest(gui=module_name):
                result = self._run_gui_pipeline_with_fakes(module_name)

                self.assertEqual(result["copy_call"].args, (result["image_paths"], result["input_root"]))
                self.assertEqual(result["crop_call"]["input_dir"], result["run_dir"])
                self.assertEqual(result["crop_call"].get("out_dir"), result["run_dir"])
                self.assertNotIn("clean", result["crop_call"])
                self.assertEqual(
                    result["scan2_call"]["model_dir"],
                    os.path.join(result["run_dir"], "stage2_fields", "model"),
                )
                self.assertEqual(
                    result["scan2_call"]["sn_dir"],
                    os.path.join(result["run_dir"], "stage2_fields", "sn"),
                )

    def test_gui_starts_ocr_prewarm_after_init(self):
        sys.modules.pop("gui_app", None)
        import gui_app

        source = inspect.getsource(gui_app.App.__init__)

        self.assertIn("start_ocr_prewarm_thread(log=self.write_log)", source)

    def test_gui_prewarm_can_be_disabled_without_importing_scan2(self):
        sys.modules.pop("gui_pipeline", None)
        sys.modules.pop("scan2", None)
        import gui_pipeline

        with mock.patch.dict(os.environ, {"HUAWEIOCR_PREWARM_OCR": "0"}):
            thread = gui_pipeline.start_ocr_prewarm_thread(log=lambda _msg: None)

        self.assertIsNone(thread)
        self.assertNotIn("scan2", sys.modules)


class OCRPrewarmTest(unittest.TestCase):
    def test_scan2_prewarm_initializes_singleton_and_runs_probe_once(self):
        scan2 = _import_scan2()
        engine = object()
        calls = []

        scan2.OCR_ENGINE = None
        scan2.OCR_PREWARM_STARTED = False
        scan2.OCR_PREWARM_DONE = False
        scan2.init_ocr = lambda: engine
        scan2.ocr_one_image = lambda ocr, img: calls.append((ocr, img)) or ([], "")

        self.assertTrue(scan2.prewarm_ocr_engine())
        self.assertIs(scan2.OCR_ENGINE, engine)
        self.assertTrue(scan2.OCR_PREWARM_DONE)
        self.assertEqual(len(calls), 1)

        self.assertTrue(scan2.prewarm_ocr_engine())
        self.assertEqual(len(calls), 1)

    def test_scan2_prewarm_failure_can_retry_later(self):
        scan2 = _import_scan2()
        calls = []

        def fail_once(_ocr, _img):
            calls.append("fail")
            raise RuntimeError("probe failed")

        scan2.OCR_ENGINE = None
        scan2.OCR_PREWARM_STARTED = False
        scan2.OCR_PREWARM_DONE = False
        scan2.ocr_one_image = fail_once

        self.assertFalse(scan2.prewarm_ocr_engine())
        self.assertFalse(scan2.OCR_PREWARM_STARTED)
        self.assertFalse(scan2.OCR_PREWARM_DONE)
        self.assertEqual(calls, ["fail"])


class BarcodeCliBudgetTest(unittest.TestCase):
    @unittest.skipUnless(os.name == "nt", "Windows process flags only")
    def test_global_subprocess_patch_hides_check_output_children(self):
        sys.modules.pop("win_subprocess", None)
        sys.modules.pop("huaweiocr.io.win_subprocess", None)
        import win_subprocess

        original_popen = subprocess.Popen
        try:
            class FakePopen:
                calls = []

                def __init__(self, *args, **kwargs):
                    self.args = args
                    self.kwargs = kwargs
                    self.returncode = 0
                    FakePopen.calls.append((args, kwargs))

                def __enter__(self):
                    return self

                def __exit__(self, *_exc_info):
                    return False

                def communicate(self, input=None, timeout=None):
                    return b"ok", b""

                def poll(self):
                    return self.returncode

                def kill(self):
                    self.returncode = -1

                def wait(self, timeout=None):
                    return self.returncode

            subprocess.Popen = FakePopen
            if hasattr(subprocess, "_huaweiocr_hidden_windows"):
                delattr(subprocess, "_huaweiocr_hidden_windows")
            win_subprocess.hide_subprocess_windows()

            self.assertTrue(issubclass(subprocess.Popen, FakePopen))

            class LibraryPopenSubclass(subprocess.Popen):
                pass

            self.assertTrue(issubclass(LibraryPopenSubclass, FakePopen))
            subprocess.check_output(["cmd.exe", "/c", "echo", "ok"])

            kwargs = FakePopen.calls[-1][1]
            self.assertTrue(kwargs["creationflags"] & subprocess.CREATE_NO_WINDOW)
            self.assertTrue(kwargs["creationflags"] & subprocess.DETACHED_PROCESS)
            self.assertEqual(kwargs["startupinfo"].wShowWindow, subprocess.SW_HIDE)
        finally:
            subprocess.Popen = original_popen
            if hasattr(subprocess, "_huaweiocr_hidden_windows"):
                delattr(subprocess, "_huaweiocr_hidden_windows")

    @unittest.skipUnless(os.name == "nt", "Windows process flags only")
    def test_run_cli_uses_detached_hidden_process_flags(self):
        barcode = _import_barcode()

        with mock.patch.object(barcode.subprocess, "run") as run_mock:
            barcode._run_cli(["BarcodeReaderCLI.exe", "--version"])

        kwargs = run_mock.call_args.kwargs
        self.assertTrue(kwargs["creationflags"] & barcode.subprocess.CREATE_NO_WINDOW)
        self.assertTrue(kwargs["creationflags"] & barcode.subprocess.DETACHED_PROCESS)
        self.assertEqual(kwargs["startupinfo"].wShowWindow, barcode.subprocess.SW_HIDE)

    def test_decode_small_patch_caps_cli_attempts(self):
        barcode = _import_barcode()
        fake_img = types.SimpleNamespace(shape=(10, 10))
        calls = []

        def fake_cli(_img, _tag, **_kwargs):
            calls.append(_tag)
            return []

        with mock.patch.object(barcode, "decode_with_transforms", return_value=[]):
            with mock.patch.object(barcode, "crop_bar_band", return_value=fake_img):
                with mock.patch.object(barcode, "enhance_band", return_value=(fake_img, fake_img)):
                    with mock.patch.object(barcode, "_rotate90", side_effect=lambda img, _k: img):
                        with mock.patch.object(barcode, "decode_with_cli", side_effect=fake_cli):
                            barcode.decode_small_patch(fake_img)

        self.assertEqual(len(calls), barcode.CLI_MAX_CALLS_PER_PATCH)


class Scan2DebugLogTest(unittest.TestCase):
    def test_info_mode_does_not_create_debug_log(self):
        scan2 = _import_scan2()
        with tempfile.TemporaryDirectory() as root:
            log_path = os.path.join(root, "debug.log")
            scan2.DEBUG_LOG_PATH = log_path
            scan2.set_log_level("info")

            scan2.append_debug("MODEL path C:/secret/customer_4E25A017ABCDEFG")

            self.assertFalse(os.path.exists(log_path))

    def test_debug_log_masks_sensitive_text(self):
        scan2 = _import_scan2()
        with tempfile.TemporaryDirectory() as root:
            log_path = os.path.join(root, "debug.log")
            scan2.DEBUG_LOG_PATH = log_path
            scan2.set_log_level("debug")

            scan2.append_sensitive_debug(r"SN C:\customer\asset_4E25A017ABCDEFG.png | 4E25A017ABCDEFG")

            with open(log_path, "r", encoding="utf-8") as f:
                data = f.read()
            self.assertNotIn("4E25A017ABCDEFG", data)
            self.assertNotIn("customer", data)
            self.assertNotIn("asset_4E25A017ABCDEFG.png", data)
            self.assertIn("4E25", data)


class AppPathsInstallTest(unittest.TestCase):
    def setUp(self):
        sys.modules.pop("app_paths", None)
        sys.modules.pop("huaweiocr.io.paths_runtime", None)
        self._real_app_paths = importlib.import_module("app_paths")

    def tearDown(self):
        sys.modules.pop("app_paths", None)
        sys.modules.pop("huaweiocr.io.paths_runtime", None)
        sys.modules["app_paths"] = self._real_app_paths

    def test_override_root_refuses_to_replace_unmarked_external_child(self):
        import app_paths

        with tempfile.TemporaryDirectory() as root:
            bundled = os.path.join(root, "bundled", "models", "official_models")
            source_model = os.path.join(bundled, "model_a")
            os.makedirs(source_model)
            with open(os.path.join(source_model, "weights.bin"), "wb") as f:
                f.write(b"complete")

            external_root = os.path.join(root, "external_models")
            target = os.path.join(external_root, "model_a")
            os.makedirs(target)
            with open(os.path.join(target, "partial.bin"), "wb") as f:
                f.write(b"partial")

            def fake_resource_path(*parts):
                return os.path.join(root, "bundled", *parts)

            with mock.patch.object(app_paths, "get_resource_path", side_effect=fake_resource_path):
                with mock.patch.dict(os.environ, {"HUAWEIOCR_MODEL_DIR": external_root}, clear=False):
                    with self.assertRaisesRegex(RuntimeError, "unmanaged model directory"):
                        app_paths.ensure_models_installed()

            self.assertTrue(os.path.exists(os.path.join(target, "partial.bin")))
            self.assertFalse(os.path.exists(os.path.join(target, "weights.bin")))
            self.assertFalse(os.path.exists(os.path.join(external_root, app_paths.MODEL_ROOT_MARKER)))

    def test_override_root_marker_allows_replacing_incomplete_child(self):
        import app_paths

        with tempfile.TemporaryDirectory() as root:
            bundled = os.path.join(root, "bundled", "models", "official_models")
            source_model = os.path.join(bundled, "model_a")
            os.makedirs(source_model)
            with open(os.path.join(source_model, "weights.bin"), "wb") as f:
                f.write(b"complete")

            external_root = os.path.join(root, "external_models")
            os.makedirs(external_root)
            with open(os.path.join(external_root, app_paths.MODEL_ROOT_MARKER), "w", encoding="utf-8") as f:
                f.write("managed\n")
            target = os.path.join(external_root, "model_a")
            os.makedirs(target)
            with open(os.path.join(target, "partial.bin"), "wb") as f:
                f.write(b"partial")

            def fake_resource_path(*parts):
                return os.path.join(root, "bundled", *parts)

            with mock.patch.object(app_paths, "get_resource_path", side_effect=fake_resource_path):
                with mock.patch.dict(os.environ, {"HUAWEIOCR_MODEL_DIR": external_root}, clear=False):
                    app_paths.ensure_models_installed()

            self.assertTrue(os.path.exists(os.path.join(target, "weights.bin")))
            self.assertTrue(os.path.exists(os.path.join(target, app_paths.MODEL_INSTALL_MARKER)))
            self.assertFalse(os.path.exists(os.path.join(target, "partial.bin")))

    def test_incomplete_model_dir_is_replaced(self):
        import app_paths

        with tempfile.TemporaryDirectory() as root:
            bundled = os.path.join(root, "bundled", "models", "official_models")
            source_model = os.path.join(bundled, "model_a")
            os.makedirs(source_model)
            with open(os.path.join(source_model, "weights.bin"), "wb") as f:
                f.write(b"complete")

            data_dir = os.path.join(root, "data")
            target = os.path.join(data_dir, "models", "official_models", "model_a")
            os.makedirs(target)
            with open(os.path.join(target, "partial.bin"), "wb") as f:
                f.write(b"partial")

            def fake_resource_path(*parts):
                return os.path.join(root, "bundled", *parts)

            with mock.patch.object(app_paths, "get_resource_path", side_effect=fake_resource_path):
                with mock.patch.dict(os.environ, {"HUAWEIOCR_DATA_DIR": data_dir, "HUAWEIOCR_MODEL_DIR": ""}, clear=False):
                    app_paths.ensure_models_installed()

            self.assertTrue(os.path.exists(os.path.join(target, "weights.bin")))
            self.assertTrue(os.path.exists(os.path.join(target, app_paths.MODEL_INSTALL_MARKER)))
            self.assertFalse(os.path.exists(os.path.join(target, "partial.bin")))

    def test_stale_model_install_lock_is_recovered(self):
        import app_paths

        with tempfile.TemporaryDirectory() as root:
            bundled = os.path.join(root, "bundled", "models", "official_models")
            source_model = os.path.join(bundled, "model_a")
            os.makedirs(source_model)
            with open(os.path.join(source_model, "weights.bin"), "wb") as f:
                f.write(b"complete")

            data_dir = os.path.join(root, "data")
            target_root = os.path.join(data_dir, "models", "official_models")
            os.makedirs(target_root)
            lock_path = os.path.join(target_root, ".huaweiocr_install.lock")
            with open(lock_path, "w", encoding="utf-8") as f:
                f.write("")
            old = time.time() - 10
            os.utime(lock_path, (old, old))

            def fake_resource_path(*parts):
                return os.path.join(root, "bundled", *parts)

            with mock.patch.object(app_paths, "get_resource_path", side_effect=fake_resource_path):
                with mock.patch.dict(os.environ, {"HUAWEIOCR_DATA_DIR": data_dir, "HUAWEIOCR_MODEL_DIR": ""}, clear=False):
                    app_paths.ensure_models_installed()

            target = os.path.join(target_root, "model_a")
            self.assertTrue(os.path.exists(os.path.join(target, "weights.bin")))
            self.assertTrue(os.path.exists(os.path.join(target, app_paths.MODEL_INSTALL_MARKER)))
            self.assertFalse(os.path.exists(lock_path))

    def test_stale_lock_reclaim_does_not_remove_changed_lock(self):
        import app_paths

        with tempfile.TemporaryDirectory() as root:
            lock_path = os.path.join(root, ".huaweiocr_install.lock")
            with open(lock_path, "w", encoding="utf-8") as f:
                f.write("")
            old = time.time() - 10
            os.utime(lock_path, (old, old))
            real_read = app_paths._read_lock_snapshot
            calls = {"count": 0}

            def racing_read(path):
                snapshot = real_read(path)
                calls["count"] += 1
                if calls["count"] == 1:
                    with open(lock_path, "w", encoding="utf-8") as f:
                        f.write(f"{os.getpid()}\n{time.time()}\n")
                return snapshot

            with mock.patch.object(app_paths, "_read_lock_snapshot", side_effect=racing_read):
                self.assertFalse(app_paths._reclaim_stale_lock(lock_path))

            self.assertTrue(os.path.exists(lock_path))
            with open(lock_path, "r", encoding="utf-8") as f:
                data = f.read()
            self.assertIn(str(os.getpid()), data)


if __name__ == "__main__":
    unittest.main()
